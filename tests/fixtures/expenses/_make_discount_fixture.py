"""Generates discount_minimal.xlsx — 5 rows on sheet 1 + 2 rows on sheet 2."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl

OUT = Path(__file__).parent / "discount_minimal.xlsx"


def main():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "עסקאות במועד החיוב"
    ws1.cell(row=1, column=1, value="כל המשתמשים (1)")
    ws1.cell(row=2, column=1, value="כל הכרטיסים (1)")
    ws1.cell(row=3, column=1, value="2026-01-01-2026-04-30")
    headers = ["תאריך עסקה", "שם בית העסק", "קטגוריה",
               "4 ספרות אחרונות של כרטיס האשראי", "סוג עסקה",
               "סכום חיוב", "מטבע חיוב", "סכום עסקה מקורי",
               "מטבע עסקה מקורי", "תאריך חיוב", "הערות", "תיוגים",
               "מועדון הנחות", "מפתח דיסקונט", "אופן ביצוע ההעסקה",
               "שער המרה ממטבע מקור/התחשבנות לש\"ח"]
    for i, h in enumerate(headers, start=1):
        ws1.cell(row=4, column=i, value=h)
    rows1 = [
        # date,         merchant,                    category,           last4, type,    chg,  ccy,  orig, orig_ccy, chg_dt,         notes,                tags, club, key, method, fx
        ("01-01-2026", "PAYBOX TEL AVIV IL",         "העברת כספים",       "2923", "רגילה",    235, "₪",  235,  "₪",      "10-01-2026", "למי: ועד בית",       None, None, None, "טלפוני",     None),
        ("05-01-2026", "פז אפליקציית יילו",          "דלק, חשמל וגז",     "2923", "רגילה",    225, "₪",  225,  "₪",      "10-01-2026", None,                  None, None, None, "טלפוני",     None),
        ("06-01-2026", "T C תאי צ'ין",               "מסעדות, קפה וברים", "2923", "רגילה",    270, "₪",  270,  "₪",      "10-01-2026", None,                  None, None, None, "טלפוני",     None),
        ("12-01-2026", "ALIEXPRESS LUXEMBOURG LU",   "עיצוב הבית",        "2923", "חיוב חודשי", -25.31, "₪", 25.31, "₪",   "10-02-2026", "ביטול עסקה",          None, None, None, "אינטרנט",   None),  # refund
        ("15-01-2026", "ויקטורי חיפה שער עי",        "מזון וצריכה",       "2923", "רגילה",    387, "₪",  387,  "₪",      "10-02-2026", None,                  None, None, None, "תשלום בנייד", None),
    ]
    for i, row in enumerate(rows1, start=5):
        for j, v in enumerate(row, start=1):
            ws1.cell(row=i, column=j, value=v)

    ws2 = wb.create_sheet(title="עסקאות חו\"ל ומט\"ח")
    ws2.cell(row=1, column=1, value="כל המשתמשים (1)")
    ws2.cell(row=2, column=1, value="כל הכרטיסים (1)")
    ws2.cell(row=3, column=1, value="2026-01-01-2026-04-30")
    for i, h in enumerate(headers, start=1):
        ws2.cell(row=4, column=i, value=h)
    rows2 = [
        ("12-02-2026", "ALIEXPRESS LUXEMBOURG LU", "עיצוב הבית", "2923",
         "חיוב חודשי", 194.23, "₪", 194.23, "₪", "10-03-2026",
         "חיוב עסקת חו\"ל בש\"ח", None, None, None, "אינטרנט", None),
        ("15-02-2026", "TEMU.COM DUBLIN 4 IE",    "עיצוב הבית", "2923",
         "חיוב חודשי", 141.32, "₪", 141.32, "₪", "10-03-2026",
         "חיוב עסקת חו\"ל בש\"ח", None, None, None, "אינטרנט", None),
    ]
    for i, row in enumerate(rows2, start=5):
        for j, v in enumerate(row, start=1):
            ws2.cell(row=i, column=j, value=v)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
