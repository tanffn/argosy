"""Generates max_minimal.xlsx with 5 rows including a refund and ענף values."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl

OUT = Path(__file__).parent / "max_minimal.xlsx"


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "לאומי לישראל 882-44745280"
    ws.cell(row=1, column=1,
            value="פירוט עסקאות לחשבון לאומי לישראל 882-44745280 02/03/2026 — 31/03/2026")
    ws.cell(row=3, column=1, value="עסקאות לחיוב ב-15/04/2026: 654.88 ₪")
    headers = ["תאריך\nעסקה", "שם בית עסק", "סכום\nעסקה", "סכום\nחיוב",
               "סוג\nעסקה", "ענף", "הערות"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)

    rows = [
        # date,                merchant,           tx,    chg,    type,        anaf,           notes
        (datetime(2026, 3, 30), "ספייס אינביידרז",   355,   355,    "רגילה",      "מסעדות",        None),
        (datetime(2026, 3, 25), "ביטוח ישיר-חיים",   142,   142,    "הוראת קבע",  "ביטוח ופיננסים", None),
        (datetime(2026, 3, 21), "WIZZ AIRGR73FH",  -2097.83, -2097.83, "זיכוי",   "תיירות",        None),
        (datetime(2026, 3, 23), "אלקטרה פאוור חשמל", 293.12, 293.12,  "הוראת קבע","ריהוט ובית",   None),
        (datetime(2026, 3, 2),  "קרן מכבי-חיוב",   133.19, 133.19, "הוראת קבע",  "רפואה ובריאות", None),
    ]
    for i, row in enumerate(rows, start=5):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
