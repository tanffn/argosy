"""Generates tests/fixtures/expenses/isracard_minimal.xlsx — 5 rows
covering the format quirks: NIS-only, USD-only, refund, standing-order."""
from __future__ import annotations

from pathlib import Path

import openpyxl

OUT = Path(__file__).parent / "isracard_minimal.xlsx"


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "פירוט עסקאות"
    # Layout (rows 1-indexed in openpyxl):
    # row 5  col 1: card header line
    # row 5  col 8: total NIS  ('₪ 250.00')
    # row 6  col 1: 'על שם אריאל יעקב'
    # row 7  col 8: 'לחיוב ב-15.04'
    # row 13 col 1..8: header row
    # rows 14+: transactions
    ws.cell(row=5, column=1, value="פלטינה מסטרקארד - 1266")
    ws.cell(row=5, column=8, value="₪ 250.00")
    ws.cell(row=6, column=1, value="על שם אריאל יעקב")
    ws.cell(row=7, column=8, value="לחיוב ב-15.04")
    headers = ["תאריך רכישה", "שם בית עסק", "סכום עסקה", "מטבע עסקה",
               "סכום חיוב", "מטבע חיוב", "מס' שובר", "פירוט נוסף"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=13, column=i, value=h)
    rows = [
        # date,        merchant,     tx_amt, tx_ccy, charge_amt, charge_ccy, voucher, extras
        ("08.04.26", "NETFLIX.COM",   69.9,  "₪",    69.9,       "₪",        "143142516", "אתר חו\"ל\nהוראת קבע"),
        ("05.04.26", "הזמנה משלוח אוכל", 131,   "₪",    127.07,     "₪",        "136908473", "הנחה ₪3.93"),
        ("24.03.26", "NAME-CHEAP.COM", 12.18, "$",    12.18,      "$",        "072314356", "אתר חו\"ל\nהוראת קבע"),
        ("22.03.26", "ZARA",          -50.0, "₪",    -50.0,      "₪",        "075759881", ""),  # refund
        ("18.03.26", "PAYPAL *VENDOR", 16.11, "₪",    16.11,      "₪",        "091130802", "אתר חו\"ל"),
    ]
    for i, row in enumerate(rows, start=14):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
