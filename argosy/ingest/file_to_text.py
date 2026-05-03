"""Convert any commonly-attached doc to plain text for downstream
intake-extractor / intake-turn agents.

Supported types (by extension + content-type sniff):
  .md / .markdown / .txt          → utf-8 decode, return as-is
  .csv / .tsv                     → utf-8 decode, return as-is
  .pdf                            → pypdf, page-by-page text join
  .xlsx                           → openpyxl, per-sheet CSV-like text

Returns a `FileToText` dataclass:
  filename: str
  content_type: str
  extracted_text: str
  warnings: list[str]      # non-fatal issues per file (e.g., empty pages)
  page_or_sheet_count: int # 0 for plain text; pages for PDF; sheets for XLSX

Raises `UnsupportedFileTypeError` for anything else.
Raises `FileTooLargeError` when bytes > 5 MB.

Pure helper — no DB, no API, no logging side effects beyond an
optional structlog `_log.warning` per warning.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO

from argosy.logging import get_logger

_log = get_logger("argosy.ingest.file_to_text")

# Hard cap. Some legitimate broker XLSX exports run a few MB; we still
# want to keep agent prompts sane, so 5 MB is the ceiling.
MAX_BYTES = 5 * 1024 * 1024  # 5 MB


class UnsupportedFileTypeError(ValueError):
    """Raised when a filename/content-type pair isn't in the whitelist."""


class FileTooLargeError(ValueError):
    """Raised when the uploaded blob exceeds MAX_BYTES."""


@dataclass
class FileToText:
    """Result of a successful conversion."""

    filename: str
    content_type: str
    extracted_text: str
    warnings: list[str] = field(default_factory=list)
    page_or_sheet_count: int = 0


# ----------------------------------------------------------------------
# Whitelist
# ----------------------------------------------------------------------

# Extensions we accept (lowercase, with dot). Each maps to a converter key.
_EXT_TO_KIND: dict[str, str] = {
    ".md": "text",
    ".markdown": "text",
    ".txt": "text",
    ".csv": "text",
    ".tsv": "text",
    ".pdf": "pdf",
    ".xlsx": "xlsx",
}

# Content types we accept as a fallback when the extension is missing or
# unhelpful. Maps to the same converter key as _EXT_TO_KIND.
_CT_TO_KIND: dict[str, str] = {
    "text/plain": "text",
    "text/markdown": "text",
    "text/x-markdown": "text",
    "text/csv": "text",
    "text/tab-separated-values": "text",
    "application/pdf": "pdf",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
}


def _resolve_kind(filename: str, content_type: str) -> str:
    """Return one of {'text','pdf','xlsx'} or raise UnsupportedFileTypeError."""
    lower = (filename or "").lower()
    for ext, kind in _EXT_TO_KIND.items():
        if lower.endswith(ext):
            return kind
    ct = (content_type or "").lower().split(";", 1)[0].strip()
    if ct in _CT_TO_KIND:
        return _CT_TO_KIND[ct]
    raise UnsupportedFileTypeError(
        f"Unsupported file type for {filename!r} "
        f"(content-type={content_type!r}). Supported: "
        ".md, .markdown, .txt, .csv, .tsv, .pdf, .xlsx"
    )


# ----------------------------------------------------------------------
# Converters
# ----------------------------------------------------------------------


def _convert_text(data: bytes) -> tuple[str, list[str], int]:
    """Plain UTF-8 decode. Returns (text, warnings, page_or_sheet_count=0)."""
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError as exc:
        # Fall back to utf-8 with replacement so we still return something
        # useful, but flag the warning so callers can show it.
        text = data.decode("utf-8", errors="replace")
        return text, [f"Non-UTF-8 bytes were replaced: {exc}"], 0
    return text, [], 0


def _convert_pdf(data: bytes) -> tuple[str, list[str], int]:
    """PDF → text (page-by-page join). Empty pages produce a warning."""
    # Import inside the function so importing this module never requires
    # the optional dep on environments where it's unavailable.
    from pypdf import PdfReader  # type: ignore[import-not-found]

    reader = PdfReader(BytesIO(data))
    pages: list[str] = []
    warnings: list[str] = []
    for idx, page in enumerate(reader.pages, start=1):
        try:
            txt = page.extract_text() or ""
        except Exception as exc:  # pragma: no cover - defensive
            txt = ""
            warnings.append(f"Page {idx}: extraction error ({exc}).")
        if not txt.strip():
            warnings.append(f"Page {idx} is empty.")
        pages.append(txt)
    text = "\n\n".join(pages)
    for w in warnings:
        _log.warning("file_to_text.pdf_warning", detail=w)
    return text, warnings, len(reader.pages)


def _convert_xlsx(data: bytes) -> tuple[str, list[str], int]:
    """XLSX → CSV-like text per sheet, separated by sheet headers.

    Each sheet is rendered as:

        ## Sheet: <sheet name>
        col_a,col_b,col_c
        v1,v2,v3
        ...

    Empty sheets produce a warning and contribute only their header.
    """
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    sections: list[str] = []
    warnings: list[str] = []

    def _cell(v: object) -> str:
        if v is None:
            return ""
        s = str(v)
        # Quote anything containing a comma or newline.
        if "," in s or "\n" in s or '"' in s:
            return '"' + s.replace('"', '""') + '"'
        return s

    for sheet in wb.worksheets:
        rows: list[str] = []
        any_rows = False
        for row in sheet.iter_rows(values_only=True):
            any_rows = True
            if all(c is None for c in row):
                continue
            rows.append(",".join(_cell(c) for c in row))
        if not any_rows or not rows:
            warnings.append(f"Sheet {sheet.title!r} is empty.")
        sections.append(f"## Sheet: {sheet.title}\n" + "\n".join(rows))
    text = "\n\n".join(sections)
    sheet_count = len(wb.worksheets)
    try:
        wb.close()
    except Exception:
        pass
    for w in warnings:
        _log.warning("file_to_text.xlsx_warning", detail=w)
    return text, warnings, sheet_count


# ----------------------------------------------------------------------
# Public entry point
# ----------------------------------------------------------------------


def convert_to_text(
    *, filename: str, content_type: str, data: bytes
) -> FileToText:
    """Convert `data` to plain text per the resolved kind.

    Raises:
      UnsupportedFileTypeError: filename/content-type isn't in the whitelist
      FileTooLargeError: len(data) > MAX_BYTES
    """
    if len(data) > MAX_BYTES:
        raise FileTooLargeError(
            f"File too large ({len(data):,} bytes; limit is {MAX_BYTES:,})."
        )
    kind = _resolve_kind(filename, content_type)
    if kind == "text":
        text, warnings, count = _convert_text(data)
    elif kind == "pdf":
        text, warnings, count = _convert_pdf(data)
    elif kind == "xlsx":
        text, warnings, count = _convert_xlsx(data)
    else:  # pragma: no cover - _resolve_kind only returns the three above
        raise UnsupportedFileTypeError(f"Unhandled kind: {kind!r}")
    return FileToText(
        filename=filename,
        content_type=content_type,
        extracted_text=text,
        warnings=warnings,
        page_or_sheet_count=count,
    )


__all__ = [
    "FileToText",
    "FileTooLargeError",
    "MAX_BYTES",
    "UnsupportedFileTypeError",
    "convert_to_text",
]
