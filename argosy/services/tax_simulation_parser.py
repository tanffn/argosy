# argosy/services/tax_simulation_parser.py
"""Parse a Schwab/ESOP simulated tax report (RSU + ESPP tabs, bilingual EN-HE headers).

Each data row is one sellable lot/tranche with its simulated Israeli tax treatment. The
load-bearing field is ``Holding Period`` = "OK" (past the Section-102 clock → capital
track ~25%) vs "Breaking" (not yet eligible → ordinary income ~62%). This turns the NVDA
deconcentration POLICY into a lot-exact, tax-aware schedule.

Bilingual headers ("Grant Award ID- מספר הענקה") are mapped to canonical keys by their
English keyword, so a re-export keeps working. Pure + dependency-light so it is unit-
testable and reusable by the upload pipeline.
"""
from __future__ import annotations

from dataclasses import dataclass, field

# English keyword (substring of the bilingual header) -> canonical field.
_HEADER_MAP = {
    "Number of Shares Requested to Sell": "shares",
    "Simulation Date": "simulation_date",
    "Grant Award ID": "grant_id",
    "Holding Period": "holding_period",
    "Grant Date": "grant_date",
    "Purchase Date": "purchase_date",
    "Employee Purchase Price": "purchase_price_usd",
    "Sale Price USD": "sale_price_usd",
    "Grant Stock Price (For Tax)": "cost_basis_usd",        # RSU tax basis
    "Share Price on Purchase Date": "purchase_close_usd",   # ESPP
    "Capital Income USD": "capital_income_usd",
    "Ordinary Income USD": "ordinary_income_usd",
    "Capital Tax Rate": "capital_tax_rate",
    "Ordinary Tax Rate": "ordinary_tax_rate",
    "Amount Wired to Bank in USD": "net_proceeds_usd",
    "Amount Wired to Bank  in ILS": "net_proceeds_ils",
}


@dataclass
class TaxSimLot:
    plan_type: str                 # "RSU" | "ESPP"
    shares: float
    holding_period: str            # "OK" | "Breaking"
    eligible: bool                 # holding_period == "OK" (capital track)
    grant_id: str = ""
    grant_date: str = ""
    purchase_date: str = ""
    sale_price_usd: float | None = None
    cost_basis_usd: float | None = None
    capital_income_usd: float | None = None
    ordinary_income_usd: float | None = None
    net_proceeds_usd: float | None = None
    simulation_date: str = ""


@dataclass
class TaxSimReport:
    simulation_date: str
    lots: list[TaxSimLot] = field(default_factory=list)

    def eligible_shares(self, plan_type: str | None = None) -> float:
        return sum(l.shares for l in self.lots
                   if l.eligible and (plan_type is None or l.plan_type == plan_type))

    def breaking_shares(self, plan_type: str | None = None) -> float:
        return sum(l.shares for l in self.lots
                   if not l.eligible and (plan_type is None or l.plan_type == plan_type))

    def eligible_by_grant(self) -> dict[str, float]:
        out: dict[str, float] = {}
        for l in self.lots:
            if l.eligible and l.grant_id:
                out[l.grant_id] = out.get(l.grant_id, 0.0) + l.shares
        return out


def _map_headers(header_row: list) -> dict[int, str]:
    """Map column index -> canonical field by matching the English keyword."""
    idx: dict[int, str] = {}
    for i, cell in enumerate(header_row):
        if not isinstance(cell, str):
            continue
        for kw, canon in _HEADER_MAP.items():
            if kw in cell:
                idx[i] = canon
                break
    return idx


def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_rows(plan_type: str, rows: list[list], header_row: list) -> list[TaxSimLot]:
    """Parse data rows for one tab. A row is a lot when it has a positive share count
    AND a holding_period (the totals/disclaimer rows lack one)."""
    colmap = _map_headers(header_row)
    out: list[TaxSimLot] = []
    for row in rows:
        rec = {colmap[i]: row[i] for i in colmap if i < len(row)}
        shares = _to_float(rec.get("shares"))
        hp = rec.get("holding_period")
        if not shares or shares <= 0 or not isinstance(hp, str) or hp.strip() == "":
            continue  # totals row / blank / disclaimer
        hp = hp.strip()
        out.append(TaxSimLot(
            plan_type=plan_type, shares=shares, holding_period=hp,
            eligible=(hp.upper() == "OK"),
            grant_id=str(rec.get("grant_id") or "").strip(),
            grant_date=str(rec.get("grant_date") or "").strip(),
            purchase_date=str(rec.get("purchase_date") or "").strip(),
            sale_price_usd=_to_float(rec.get("sale_price_usd")),
            cost_basis_usd=_to_float(rec.get("cost_basis_usd"))
                or _to_float(rec.get("purchase_price_usd")),
            capital_income_usd=_to_float(rec.get("capital_income_usd")),
            ordinary_income_usd=_to_float(rec.get("ordinary_income_usd")),
            net_proceeds_usd=_to_float(rec.get("net_proceeds_usd")),
            simulation_date=str(rec.get("simulation_date") or "").strip(),
        ))
    return out


def parse_workbook(path: str) -> TaxSimReport:
    """Parse the xlsx at ``path`` (RSU + ESPP tabs). Header is on the row that contains
    'Number of Shares Requested to Sell'."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    lots: list[TaxSimLot] = []
    sim_date = ""
    try:
        for sn in wb.sheetnames:
            plan_type = sn.strip().upper()
            if plan_type not in ("RSU", "ESPP"):
                continue
            allrows = [list(r) for r in wb[sn].iter_rows(values_only=True)]
            header_i = next(
                (i for i, r in enumerate(allrows)
                 if any(isinstance(c, str) and "Number of Shares Requested to Sell" in c
                        for c in r)),
                None,
            )
            if header_i is None:
                continue
            header_row = allrows[header_i]
            data = allrows[header_i + 1:]
            tab_lots = parse_rows(plan_type, data, header_row)
            lots.extend(tab_lots)
            if not sim_date:
                sim_date = next((l.simulation_date for l in tab_lots if l.simulation_date), "")
    finally:
        wb.close()
    return TaxSimReport(simulation_date=sim_date, lots=lots)
