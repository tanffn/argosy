"""NVIDIA RSU/ESPP simulation-report parser — the per-grant §102 tax authority.

The Schwab ESOP "Nvidia simulation Report.xlsx" is the *authoritative* per-grant
Israeli §102 tax model: for every grant it states the capital/ordinary income
split, the tax rates, the advance tax deducted, and the net "Amount Wired to
Bank". This module:

  1. Parses the ``RSU`` (and optionally ``ESPP``) sheet into per-row records and
     aggregates them per grant.
  2. Exposes the *exact* §102 formula reverse-engineered from the sheet (verified
     to 0.0 USD error against every data row), so net proceeds can be computed
     for the user's ACTUAL sale quantities — not just the sim's quantities.

§102 trustee-track model (verified against the sheet, 0.0 USD residual)
-----------------------------------------------------------------------
For a *capital-track* grant (Holding Period == 'OK' — the 2-year §102 trustee
period was met):

    ordinary_income = grant_price * shares - fees
    capital_income  = (sale_price - grant_price) * shares
    advance_tax     = capital_income * capital_rate(0.25)
                    + ordinary_income * ordinary_rate(0.6217)
    net_wired       = gross_after_fees - advance_tax - wire_fee

where ``grant_price`` is the sheet's "Grant Stock Price (For Tax)" — the
§102 ordinary-income base (the trailing-average grant value), NOT the
acquisition FMV.

For a *breaking* grant (Holding Period == 'Breaking' — the 2-year period was
NOT met): the entire proceeds are ordinary income:

    ordinary_income = gross_after_fees
    capital_income  = 0
    advance_tax     = ordinary_income * ordinary_rate(0.6217)

``effective_retention = net_wired / gross_after_fees`` is therefore strongly
grant-dependent: ~0.72 for old capital-heavy grants (low grant_price → large
capital portion @25%), down to ~0.38 for recent breaking grants (100% ordinary
@62%). NEVER assume a flat rate.

Read-only, pure-Python; ``Decimal`` would be ideal but the sheet itself is
``float``, so we keep ``float`` and assert tolerances at the test boundary.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Column layout (0-indexed) of the RSU sheet. Header is Excel row 3
# (iter index 2); data rows follow until a totals row with empty grant id.
# ---------------------------------------------------------------------------

_COL_SHARES = 0          # Number of Shares Requested to Sell
_COL_GRANT_ID = 2        # Grant Award ID
_COL_HOLDING = 3         # Holding Period ('OK' | 'Breaking')
_COL_GRANT_DATE = 4      # Grant Date (DD/MM/YYYY)
_COL_SALE_PRICE = 5      # Sale Price USD
_COL_BROKER_FEE = 6      # Broker Fee USD
_COL_TRUSTEE_FEE = 7     # Trustee Fee USD
_COL_GRANT_PRICE = 8     # Grant Stock Price (For Tax) USD — the §102 ordinary base
_COL_GROSS = 9           # Gross Proceeds After Fees USD
_COL_FX = 11             # Exchange Rate USD -> ILS
_COL_CAP_RATE = 12       # Capital Tax Rate (0.25)
_COL_CAP_INCOME = 13     # Capital Income USD
_COL_ORD_RATE = 15       # Ordinary Tax Rate (0.6217)
_COL_ORD_INCOME = 16     # Ordinary Income USD
_COL_ADV_TAX = 18        # Total Advance Tax Deducted USD
_COL_WIRE_FEE = 20       # Wire Fee USD
_COL_WIRED = 22          # Amount Wired to Bank USD (the NET — ground truth)

# ESPP sheet is laid out the same up to a column shift (it has Employee
# Purchase Price + Purchase Date instead of Grant Award ID + Grant Date, so
# the income/tax columns sit at the same indices). We parse it generically.
_ESPP_COL_SHARES = 0
_ESPP_COL_HOLDING = 2
_ESPP_COL_PURCHASE_PRICE = 3
_ESPP_COL_SALE_PRICE = 5
_ESPP_COL_BROKER_FEE = 6
_ESPP_COL_TRUSTEE_FEE = 7
_ESPP_COL_GROSS = 9
_ESPP_COL_CAP_RATE = 12
_ESPP_COL_CAP_INCOME = 13
_ESPP_COL_ORD_RATE = 15
_ESPP_COL_ORD_INCOME = 16
_ESPP_COL_ADV_TAX = 18
_ESPP_COL_WIRE_FEE = 20
_ESPP_COL_WIRED = 22


HOLDING_CAPITAL = "OK"
HOLDING_BREAKING = "Breaking"

# The sim's ordinary withholding rate (0.6217) reproduces the sim's "Amount
# Wired" exactly, but OVER-withholds vs the actual ESOP bank wire. The actual
# wires fit an effective ordinary rate of ~0.50 (Israeli top marginal income
# tax 47% + 3% surtax). The extra ~12.17% the sim withholds is employee
# National Insurance / health (Bituah Leumi, full band — capped) reconciled
# via payroll, NOT deducted at the ESOP wire. Use this rate to reconcile
# against REAL Leumi inflows (codex-validated: closes the residual from ~1.5%
# to <0.05% across all 5 real 2026 sales). NOTE: this reconciles the ESOP CASH
# WIRE, not the final annual tax return (which can still adjust NI/health caps,
# annual surtax, FX/NIS basis, and credits). See
# tmp_review/rsu_recon_codex_verdict_r2.md.
WIRE_ORDINARY_RATE = 0.50


@dataclass(frozen=True)
class GrantTaxModel:
    """Per-grant §102 tax model derived from the simulation sheet.

    All monetary fields are the sheet's *aggregate over that grant's rows* (the
    sim breaks a grant across several quantity rows). The rates + grant_price
    are constant per grant. ``effective_retention`` and ``capital_fraction``
    summarise the split; ``net_for_shares`` recomputes net for an arbitrary
    quantity using the verified §102 formula.
    """

    grant_id: str
    holding_period: str               # 'OK' (capital track) | 'Breaking'
    grant_date: date | None
    grant_price_usd: float            # §102 ordinary-income base ("Grant Stock Price For Tax")
    sale_price_usd: float             # sim sale price (illustrative; ~$204.65)
    capital_rate: float               # 0.25
    ordinary_rate: float              # 0.6217
    shares: int                       # total sim shares for this grant
    gross_after_fees_usd: float
    capital_income_usd: float
    ordinary_income_usd: float
    advance_tax_usd: float
    wire_fee_usd: float
    net_wired_usd: float
    fx_usd_ils: float

    @property
    def is_capital_track(self) -> bool:
        return self.holding_period == HOLDING_CAPITAL

    @property
    def effective_retention(self) -> float:
        """net_wired / gross_after_fees for the sim quantity (wire fee folded
        in). Grant-dependent; this is the calibration ratio to apply to an
        actual sale of the same grant when a per-lot recompute isn't possible.
        """
        if self.gross_after_fees_usd <= 0:
            return 0.0
        return self.net_wired_usd / self.gross_after_fees_usd

    @property
    def capital_fraction(self) -> float:
        """Capital income as a fraction of total taxable income."""
        total = self.capital_income_usd + self.ordinary_income_usd
        if total <= 0:
            return 0.0
        return self.capital_income_usd / total

    def net_for_shares(
        self,
        shares: float,
        sale_price_usd: float,
        *,
        fees_usd: float = 0.0,
        wire_fee_usd: float = 0.0,
        ordinary_rate: float | None = None,
    ) -> "SaleNet":
        """Compute net proceeds for ``shares`` of THIS grant sold at
        ``sale_price_usd``, applying this grant's §102 treatment.

        ``fees_usd`` are the broker+trustee fees on the actual sale (so the
        gross-after-fees matches the broker's). ``wire_fee_usd`` is the wire
        transfer fee (one-time per wire, default 0).

        ``ordinary_rate`` overrides the grant's sim ordinary rate (0.6217). The
        sim WITHHOLDS at 0.6217, but the *actual* ESOP wire reflects an
        effective ordinary withholding of ~0.50 (Israeli top marginal 47% +
        3% surtax = 50%); the extra ~12.17% the sim withholds is employee
        National Insurance / health (capped) reconciled in payroll, not deducted
        at the wire. Pass :data:`WIRE_ORDINARY_RATE` to match actual bank
        inflows; leave ``None`` to reproduce the sim's own (conservative) net.

        Capital-track ('OK'):
            ordinary = grant_price*shares - fees
            capital  = (sale_price - grant_price)*shares
        Breaking:
            ordinary = gross_after_fees ; capital = 0
        """
        ord_rate = self.ordinary_rate if ordinary_rate is None else ordinary_rate
        gross_after_fees = sale_price_usd * shares - fees_usd
        if self.is_capital_track:
            ordinary_income = self.grant_price_usd * shares - fees_usd
            capital_income = (sale_price_usd - self.grant_price_usd) * shares
        else:
            ordinary_income = gross_after_fees
            capital_income = 0.0
        advance_tax = (
            capital_income * self.capital_rate
            + ordinary_income * ord_rate
        )
        net = gross_after_fees - advance_tax - wire_fee_usd
        return SaleNet(
            grant_id=self.grant_id,
            holding_period=self.holding_period,
            shares=shares,
            sale_price_usd=sale_price_usd,
            grant_price_usd=self.grant_price_usd,
            gross_after_fees_usd=round(gross_after_fees, 2),
            capital_income_usd=round(capital_income, 2),
            ordinary_income_usd=round(ordinary_income, 2),
            capital_rate=self.capital_rate,
            ordinary_rate=ord_rate,
            advance_tax_usd=round(advance_tax, 2),
            wire_fee_usd=round(wire_fee_usd, 2),
            net_usd=round(net, 2),
        )


@dataclass(frozen=True)
class SaleNet:
    """Result of applying a grant's §102 model to a (shares, price) sale."""

    grant_id: str
    holding_period: str
    shares: float
    sale_price_usd: float
    grant_price_usd: float
    gross_after_fees_usd: float
    capital_income_usd: float
    ordinary_income_usd: float
    capital_rate: float
    ordinary_rate: float
    advance_tax_usd: float
    wire_fee_usd: float
    net_usd: float

    @property
    def effective_retention(self) -> float:
        if self.gross_after_fees_usd <= 0:
            return 0.0
        return self.net_usd / self.gross_after_fees_usd

    @property
    def capital_fraction(self) -> float:
        total = self.capital_income_usd + self.ordinary_income_usd
        if total <= 0:
            return 0.0
        return self.capital_income_usd / total


@dataclass
class SimTaxModel:
    """The whole simulation report: per-grant §102 models + ESPP pool."""

    grants: dict[str, GrantTaxModel] = field(default_factory=dict)
    espp: GrantTaxModel | None = None
    simulation_date: date | None = None

    def grant(self, grant_id: str) -> GrantTaxModel | None:
        return self.grants.get(str(grant_id))

    @property
    def total_gross_usd(self) -> float:
        return sum(g.gross_after_fees_usd for g in self.grants.values())

    @property
    def total_net_usd(self) -> float:
        return sum(g.net_wired_usd for g in self.grants.values())


def _d(s) -> date | None:
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(s).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _num(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_sim_report(
    path: Path,
    *,
    include_espp: bool = True,
) -> SimTaxModel:
    """Parse the NVIDIA simulation report (.xlsx) into per-grant §102 models.

    The RSU sheet header is Excel row 3; data rows run until a totals row whose
    grant-id cell is empty. Each grant typically spans several quantity rows
    (the wire fee appears on only the first); we aggregate per grant.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    model = SimTaxModel()

    # --- RSU sheet ---------------------------------------------------------
    ws = wb["RSU"]
    agg: dict[str, dict] = {}
    sim_date: date | None = None
    for row in ws.iter_rows(min_row=4, values_only=True):  # data starts at Excel row 4
        if not row:
            continue
        shares = _num(row[_COL_SHARES]) if len(row) > _COL_SHARES else None
        grant_id = row[_COL_GRANT_ID] if len(row) > _COL_GRANT_ID else None
        if shares is None or grant_id in (None, ""):
            continue  # totals / disclaimer / blank row
        grant_id = str(grant_id).strip()
        if sim_date is None and len(row) > 1:
            sim_date = _d(row[1])
        a = agg.setdefault(grant_id, dict(
            shares=0.0, gross=0.0, cap=0.0, ord=0.0, adv=0.0,
            wire=0.0, wired=0.0,
        ))
        a["shares"] += shares
        a["gross"] += _num(row[_COL_GROSS]) or 0.0
        a["cap"] += _num(row[_COL_CAP_INCOME]) or 0.0
        a["ord"] += _num(row[_COL_ORD_INCOME]) or 0.0
        a["adv"] += _num(row[_COL_ADV_TAX]) or 0.0
        a["wire"] += _num(row[_COL_WIRE_FEE]) or 0.0
        a["wired"] += _num(row[_COL_WIRED]) or 0.0
        # constants (last write wins; identical across a grant's rows)
        a["holding"] = (row[_COL_HOLDING] or "").strip()
        a["grant_date"] = _d(row[_COL_GRANT_DATE])
        a["grant_price"] = _num(row[_COL_GRANT_PRICE])
        a["sale_price"] = _num(row[_COL_SALE_PRICE])
        a["cap_rate"] = _num(row[_COL_CAP_RATE]) or 0.25
        a["ord_rate"] = _num(row[_COL_ORD_RATE]) or 0.6217
        a["fx"] = _num(row[_COL_FX]) or 0.0

    model.simulation_date = sim_date
    for gid, a in agg.items():
        model.grants[gid] = GrantTaxModel(
            grant_id=gid,
            holding_period=a.get("holding", ""),
            grant_date=a.get("grant_date"),
            grant_price_usd=a.get("grant_price") or 0.0,
            sale_price_usd=a.get("sale_price") or 0.0,
            capital_rate=a.get("cap_rate", 0.25),
            ordinary_rate=a.get("ord_rate", 0.6217),
            shares=int(round(a["shares"])),
            gross_after_fees_usd=round(a["gross"], 2),
            capital_income_usd=round(a["cap"], 2),
            ordinary_income_usd=round(a["ord"], 2),
            advance_tax_usd=round(a["adv"], 2),
            wire_fee_usd=round(a["wire"], 2),
            net_wired_usd=round(a["wired"], 2),
            fx_usd_ils=a.get("fx", 0.0),
        )

    # --- ESPP sheet (optional, single pool) --------------------------------
    if include_espp and "ESPP" in wb.sheetnames:
        wse = wb["ESPP"]
        e = dict(shares=0.0, gross=0.0, cap=0.0, ord=0.0, adv=0.0,
                 wire=0.0, wired=0.0, cap_rate=0.25, ord_rate=0.6217)
        any_row = False
        for row in wse.iter_rows(min_row=4, values_only=True):
            if not row:
                continue
            sh = _num(row[_ESPP_COL_SHARES]) if len(row) > _ESPP_COL_SHARES else None
            gross = _num(row[_ESPP_COL_GROSS]) if len(row) > _ESPP_COL_GROSS else None
            if sh is None or gross is None:
                continue
            any_row = True
            e["shares"] += sh
            e["gross"] += gross
            e["cap"] += _num(row[_ESPP_COL_CAP_INCOME]) or 0.0
            e["ord"] += _num(row[_ESPP_COL_ORD_INCOME]) or 0.0
            e["adv"] += _num(row[_ESPP_COL_ADV_TAX]) or 0.0
            e["wire"] += _num(row[_ESPP_COL_WIRE_FEE]) or 0.0
            e["wired"] += _num(row[_ESPP_COL_WIRED]) or 0.0
            e["cap_rate"] = _num(row[_ESPP_COL_CAP_RATE]) or 0.25
            e["ord_rate"] = _num(row[_ESPP_COL_ORD_RATE]) or 0.6217
        if any_row and e["shares"] > 0:
            model.espp = GrantTaxModel(
                grant_id="ESPP",
                holding_period="mixed",
                grant_date=None,
                grant_price_usd=0.0,
                sale_price_usd=0.0,
                capital_rate=e["cap_rate"],
                ordinary_rate=e["ord_rate"],
                shares=int(round(e["shares"])),
                gross_after_fees_usd=round(e["gross"], 2),
                capital_income_usd=round(e["cap"], 2),
                ordinary_income_usd=round(e["ord"], 2),
                advance_tax_usd=round(e["adv"], 2),
                wire_fee_usd=round(e["wire"], 2),
                net_wired_usd=round(e["wired"], 2),
                fx_usd_ils=0.0,
            )

    wb.close()
    return model


__all__ = [
    "GrantTaxModel",
    "SaleNet",
    "SimTaxModel",
    "parse_sim_report",
    "HOLDING_CAPITAL",
    "HOLDING_BREAKING",
    "WIRE_ORDINARY_RATE",
]
